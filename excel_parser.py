#!/usr/bin/env python3
"""
excel_parser.py
---------------
Reads all XLSX files from a ./specs/ folder and outputs:
  - cameras.json
  - recorders.json

Usage:
    python3 excel_parser.py
    python3 excel_parser.py --specs-dir ./specs --output-dir .
"""

import argparse
import json
import os
import re
import sys
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("Missing dependency. Run: pip install openpyxl")
    sys.exit(1)


# ── ROW LABEL → FIELD NAME ───────────────────────────────────────────────────

CAMERA_ROW_MAP = {
    # ── Identity ──────────────────────────────────
    'Model Name':                       '__model__',
    'Status':                           'status',
    'Title':                            'title',
    'Release Date':                     'releaseDate',

    # ── Imaging ───────────────────────────────────
    'Image Sensor':                     'imageSensor',
    'Max. Resolution':                  'resolution',
    'Max Resolution':                   'resolution',
    'Resolution':                       'resolution',

    # ── Lens ──────────────────────────────────────
    'Lens Type':                        'lensType',
    'Focal Length':                     'focalLength',
    'Angular Field of View':            'fov',
    'Field of View':                    'fov',

    # ── Performance ───────────────────────────────
    'Min. Illumination':                'minIllumination',
    'Dynamic Range':                    'wdr',          # "120dB (True WDR)"
    'WDR':                              'wdr',
    'Day and Night':                    'dayNight',
    'Day & Night':                      'dayNight',

    # ── Video ─────────────────────────────────────
    'Video Compression':                'compression',
    'Max. Frame Rate':                  'maxFrameRate',
    'Max Frame Rate':                   'maxFrameRate',
    'Intelligent Video':                'videoAnalytics',  # VA description
    'Video Analytics':                  'videoAnalytics',

    # ── Audio / Alarm ─────────────────────────────
    'Two-way Audio':                    'twoWayAudio',
    'Audio In / Out':                   'audioInOut',
    'Audio In/Out':                     'audioInOut',
    'Alarm In / Out':                   'alarmInOut',
    'Alarm In/Out':                     'alarmInOut',

    # ── Storage ───────────────────────────────────
    'Edge Storage':                     'edgeStorage',

    # ── Physical ──────────────────────────────────
    'Vandal Proof Casing':              'vandal',
    'Vandal':                           'vandal',
    'Outdoor Ready':                    'outdoor',
    'Outdoor':                          'outdoor',
    'Power Source':                     'powerSource',
    'Power Consumption':                'powerConsumption',
    'Dimensions':                       'dimensions',
    'Weight':                           'weight',

    # ── Certification ─────────────────────────────
    'Approval':                         'approval',
    'Certificate':                      'approval',

    # ── Key info ──────────────────────────────────
    'Key Feature':                      'keyFeature',
    'Key Features':                     'keyFeature',

    # ── IR ────────────────────────────────────────
    'IR Distance (LEDs) - OVERSEA':     'irDistance',
    'IR Distance (LEDs) - DOMESTIC':    '__ir_domestic__',  # fallback
    'IR Distance':                      'irDistance',

    # ── Security (for FIPS detection) ─────────────
    'Security':                         '__security__',
}

RECORDER_ROW_MAP = {
    # ── Identity ──────────────────────────
    'Model Name':               '__model__',
    'Status':                   'status',
    'Release Date':             'releaseDate',
    'Title':                    'title',

    # ── Video / Channel ───────────────────
    'Video Inputs':             '__channels__',
    'IP Channels':              '__channels__',
    'Max. IP Channels':         '__channels__',
    'Max IP Channels':          '__channels__',
    'Max Channels':             '__channels__',
    'Channels':                 '__channels__',

    'Max. Incoming Throughput': 'maxBandwidth',
    'Max Incoming Throughput':  'maxBandwidth',
    'Max Bandwidth':            'maxBandwidth',
    'Max. Bandwidth':           'maxBandwidth',
    'Incoming Bandwidth':       'maxBandwidth',

    'Live + Recording + Remote':'liveRecRemote',

    'Supported Camera':         'supportedCam',
    'Supported Cameras':        'supportedCam',

    'Video Outputs':            'videoOut',
    'Video Output':             'videoOut',
    'Video Out':                'videoOut',

    'Display Layout':           'displayLayout',
    'Display Resolution':       'displayResolution',
    'Display Speed':            'displaySpeed',
    'Max. Display Throughput':  'displayThroughput',
    'Max Display Throughput':   'displayThroughput',
    'Digital Zoom':             'digitalZoom',

    'NVR Side Dewarping':       '__dewarping__',
    'Dewarping':                '__dewarping__',
    'Fisheye Dewarping':        '__dewarping__',

    # ── Recording ─────────────────────────
    'Max. Throughput':          'recBandwidth',
    'Max Throughput':           'recBandwidth',
    'Recording Bandwidth':      'recBandwidth',
    'Max Recording Bandwidth':  'recBandwidth',

    'Recording Resolution':     'recRes',
    'Max Recording Res.':       'recRes',
    'Max. Recording Res.':      'recRes',
    'Max Recording Resolution': 'recRes',

    'Recording Bitrate':        'recordingBitrate',
    'Video Compression':        'compression',
    'Recording Mode':           'recordingMode',
    'Trigger Event':            'triggerEvent',
    'Event Action':             'eventAction',

    # ── Playback ──────────────────────────
    'Performance':              'playbackPerformance',
    'Search Mode':              'searchMode',

    # ── Storage ───────────────────────────
    'HDD':                      'hdd',
    'Total Capacity':           'totalCapacity',
    'Max Total Capacity':       'totalCapacity',
    'Max. Total Capacity':      'totalCapacity',

    # ── Interface / General ───────────────
    'Two-way Audio':            'twoWayAudio',
    'Audio In / Out':           'audioInOut',
    'Audio In/Out':             'audioInOut',
    'Alarm In / Out':           'alarmInOut',
    'Alarm In/Out':             'alarmInOut',
    'Approval':                 'approval',
    'Key Feature':              'keyFeature',

    # ── Optional explicit flags ───────────
    'RAID':                     '__raid__',
    'ONVIF':                    '__onvif__',
    '4K':                       '__4k__',
    '4K UHD':                   '__4k__',
    '4K Support':               '__4k__',
}

STATUS_MAP = {
    '출시완료': '출시완료',
    '개발중':   '개발중',
    '단종':     '단종',
    'Released':     '출시완료',
    'Active':       '출시완료',
    'EOL':          '단종',
    'Discontinued': '단종',
    'Development':  '개발중',
}

# Filename keyword → category
CAM_CATEGORY_MAP = {
    'DOME':        'Dome',
    'BULLET':      'Bullet',
    'PTZ':         'PTZ',
    'FISHEYE':     'Fisheye',
    'BOX':         'Box',
    'MODULAR':     'Modular',
    'INTERCOM':    'Intercom',
    'MULTIIMAGER': 'Multi-imager',
    'MULTI':       'Multi-imager',
    'MOBILE':      'Mobile',
    'PARKING':     'Parking',
    '주차':        'Parking',
}

# Filename keyword → recorder series
REC_SERIES_MAP = {
    'DR1': 'DR1',
    'DR2': 'DR2',
    'DR6': 'DR6',
    'DR8': 'DR8',
    'IR000': 'IR-SVR',
    'IR0000': 'IR-SVR',
    'IS0000': 'IR-WS',
    'IR_IR': 'IR-SVR',
    'IR_IS': 'IR-WS',
}


# ── HELPERS ───────────────────────────────────────────────────────────────────

def is_camera_file(filename: str) -> bool:
    upper = filename.upper()
    return upper.startswith('IPCAM_') or upper.startswith('IP-CAM_')


def is_recorder_file(filename: str) -> bool:
    return filename.upper().startswith('NVR_')


def guess_camera_category(filename: str) -> str:
    # Normalize: remove hyphens and underscores for matching
    upper = filename.upper().replace('-', '').replace('_', '')
    for key, cat in CAM_CATEGORY_MAP.items():
        if key.upper().replace('-', '').replace('_', '') in upper:
            return cat
    return 'Other'


def guess_recorder_series(filename: str) -> str:
    upper = filename.upper()
    # Try exact matches first (longer keys first to avoid partial matches)
    for key in sorted(REC_SERIES_MAP.keys(), key=len, reverse=True):
        if key.upper() in upper:
            return REC_SERIES_MAP[key]
    # Fallback: look for DRn pattern
    m = re.search(r'DR(\d)', upper)
    if m:
        return f'DR{m.group(1)}'
    return 'Unknown'


def cell_str(val) -> str:
    """Convert cell value to clean string."""
    if val is None:
        return '-'
    s = str(val).strip()
    return s if s else '-'


def parse_mp(resolution_str: str) -> float:
    m = re.search(r'(\d+)\s*[xX×]\s*(\d+)', str(resolution_str or ''))
    if m:
        return round(int(m.group(1)) * int(m.group(2)) / 1_000_000, 1)
    return 0.0


def parse_lens(lenstype_str: str):
    """Returns (lensType, lensTypeTag)."""
    s = str(lenstype_str or '').lower()
    if 'motorized' in s or 'vari-focal' in s or 'varifocal' in s:
        return lenstype_str, 'motorized'
    if 'af zoom' in s or 'zoom lens' in s:
        return lenstype_str, 'motorized'
    if 'c/cs' in s or 'cs mount' in s or 'c mount' in s:
        return lenstype_str, 'cs_mount'
    return lenstype_str, 'fixed'


def get_fl_subtag(focal_length: str, lens_tag: str) -> str | None:
    fl = str(focal_length or '').strip()
    if not fl or fl == '-':
        return None
    if lens_tag == 'fixed':
        try:
            val = float(re.search(r'[\d.]+', fl).group())
        except (AttributeError, ValueError):
            return fl
        if val <= 2.0:  return '≤2.0mm'
        if val <= 2.9:  return '2.8mm'
        if val <= 3.5:  return '3.3mm'
        if val <= 4.5:  return '4.0–4.3mm'
        return '6.0mm+'
    return None  # motorized uses getMotoGroup() in JS


def detect_analytics(video_analytics: str, key_features: str) -> str:
    va = str(video_analytics or '').lower()
    kf = str(key_features or '').lower()
    if any(x in va for x in ['idla pro', 'a-cut', 'crowd detection', 'fall detection', 'abandoned']):
        return 'edgeai_plus'
    if 'edgeai+' in kf or 'edgeai plus' in kf:
        return 'edgeai_plus'
    if any(x in va for x in ['idla', 'object detection', 'intrusion', 'loitering', 'line crossing']):
        return 'edgeai'
    return 'none'

def detect_analytics_features(video_analytics: str, key_features: str = '') -> list:
    """
    Extract individual VA features from Video Analytics / Key Feature text.
    Returns display labels used by VA Selector.
    """
    text = f"{video_analytics or ''} {key_features or ''}".lower()

    # Normalize common separators / typo variations
    text = text.replace('-', ' ')
    text = text.replace('_', ' ')
    text = re.sub(r'\s+', ' ', text)

    feature_patterns = [
        {
            'label': 'A-Cut',
            'patterns': [
                r'\ba\s*cut\b',
                r'\bacut\b',
            ],
        },
        {
            'label': 'Object Detection',
            'patterns': [
                r'\bobject detection\b',
            ],
        },
        {
            'label': 'Intrusion',
            'patterns': [
                r'\bintrusion\b',
            ],
        },
        {
            'label': 'Loitering',
            'patterns': [
                r'\bloitering\b',
                r'\bloitering detection\b',
            ],
        },
        {
            'label': 'Line Crossing',
            'patterns': [
                r'\bline crossing\b',
                r'\bline crossing detection\b',
            ],
        },
        {
            'label': 'Face Detection',
            'patterns': [
                r'\bface detection\b',
            ],
        },
        {
            'label': 'Crowd Detection',
            'patterns': [
                r'\bcrowd detection\b',
            ],
        },
        {
            'label': 'Abandoned Object Detection',
            'patterns': [
                r'\babandoned object detection\b',
                r'\babandoned detection\b',
                r'\babandoned\b',
            ],
        },
        {
            'label': 'Removed Object Detection',
            'patterns': [
                r'\bremoved object detection\b',
                r'\bremoved detection\b',
                r'\bremoved\b',
            ],
        },
        {
            'label': 'Fall Detection',
            'patterns': [
                r'\bfall detection\b',
            ],
        },
    ]

    features = []

    for feature in feature_patterns:
        if any(re.search(pattern, text) for pattern in feature['patterns']):
            features.append(feature['label'])

    return features

def detect_camera_booleans(cam: dict) -> dict:
    va       = str(cam.get('videoAnalytics', '')).lower()
    approval = str(cam.get('approval', '')).upper()
    security = str(cam.get('__security__', '')).lower()
    edge     = str(cam.get('edgeStorage', '')).lower()
    audio_io = str(cam.get('audioInOut', '')).lower()
    two_way  = str(cam.get('twoWayAudio', '')).lower()
    vandal   = str(cam.get('vandal', '')).lower()
    outdoor  = str(cam.get('outdoor', '')).lower()
    alarm    = str(cam.get('alarmInOut', '')).lower()
    ir       = str(cam.get('irDistance', '')).lower()
    kf       = str(cam.get('keyFeature', '')).lower()
    lens_tag = cam.get('lensTypeTag', 'fixed')
    tier     = cam.get('analyticsTier', 'none')

    def has_value(s):
        return bool(s and s not in ['-', 'none', ''])

    cam['hasIR']           = has_value(ir) and 'warm' not in ir
    cam['hasAudio']        = has_value(audio_io) and '/' in audio_io or two_way in ('yes', 'true')
    cam['hasBuiltinMic']   = 'built in mic' in audio_io or 'built-in mic' in audio_io or 'built in mic' in kf
    cam['hasAlarm']        = has_value(alarm) and '/' in alarm
    cam['hasSD']           = has_value(edge)
    cam['hasWDR']          = tier in ('edgeai', 'edgeai_plus') or '120db' in str(cam.get('wdr', '')).lower()
    cam['hasIDLA']         = 'idla' in va or tier in ('edgeai', 'edgeai_plus')
    cam['hasVandal']       = has_value(vandal)
    cam['hasOutdoor']      = has_value(outdoor)
    cam['hasMotorizedZoom']= lens_tag == 'motorized'
    cam['isDirectIP']      = True
    cam['isONVIF']         = True
    cam['onvifProfiles']   = ['S', 'T', 'G', 'M']
    cam['isLPR']           = 'license plate' in va or 'lpr' in va or 'lpr' in kf
    cam['isActiveDeterrent']= 'active deterrence' in kf or 'warm light' in ir or 'warning light' in kf
    cam['isPano180']       = '180' in str(cam.get('fov', '')).lower() and 'pano' in str(cam.get('title', '')).lower()
    cam['hasUL']           = 'UL' in approval
    cam['hasFIPS']         = 'fips' in security

    return cam


def bool_cell(val) -> bool:
    s = str(val or '').lower().strip()
    return s in ('yes', 'true', '1', 'o', '●', '■', 'supported', 'support')

def clean_channel_text(value: str) -> str:
    """
    '16 IP Channels' -> '16'
    """
    m = re.search(r'(\d+)', str(value or ''))
    return m.group(1) if m else '-'


def clean_rec_resolution(value: str) -> str:
    """
    'Up to 12MP (Depending on IP Camera)' -> 'Up to 12MP'
    """
    s = cell_str(value)
    if s == '-':
        return '-'
    s = re.sub(r'\s*\([^)]*\)\s*', '', s).strip()
    return s if s else '-'


def clean_total_capacity(value: str) -> str:
    """
    '144TB=10TB x 8(Internal) + ...' -> '144TB'
    """
    s = cell_str(value)
    if s == '-':
        return '-'

    if '=' in s:
        return s.split('=', 1)[0].strip()

    m = re.search(r'(\d+(?:\.\d+)?\s*(?:TB|GB))', s, re.I)
    if m:
        return m.group(1).replace(' ', '')

    return s


def text_has_raid(*values) -> bool:
    text = ' '.join(str(v or '') for v in values).lower()
    return 'raid' in text

def clean_bandwidth(value: str) -> str:
    """
    '180Mbps, 480ips@UHD' -> '180Mbps'
    '320Mbps' -> '320Mbps'
    """
    s = cell_str(value)
    if s == '-':
        return '-'

    m = re.search(r'(\d+(?:\.\d+)?\s*Mbps)', s, re.I)
    if m:
        return m.group(1).replace(' ', '')

    return s.split(',', 1)[0].strip()

def text_has_4k(*values) -> bool:
    text = ' '.join(str(v or '') for v in values).lower()
    return (
        '4k' in text
        or 'uhd' in text
        or '3840' in text
        or '2160' in text
    )

def parse_transposed_xlsx(ws, row_map: dict) -> list:
    """
    Parse IDIS-style transposed XLSX:
      Row 1: col A = spec label, col B+ = model values
    Returns list of dicts.
    """
    # Build row_label → {col: value} mapping
    label_data = {}
    for row in ws.iter_rows():
        label_raw = cell_str(row[0].value)
        if label_raw == '-':
            continue
        label_data[label_raw] = {i: cell.value for i, cell in enumerate(row)}

    # Model names are in row 1, col 1+
    model_row_raw = None
    for label_raw in label_data:
        if label_raw == 'Model Name':
            model_row_raw = label_data[label_raw]
            break
    if not model_row_raw:
        return []

    # Collect model columns (skip col 0 which is the label)
    model_cols = [(col, cell_str(val)) for col, val in model_row_raw.items()
                  if col > 0 and val and cell_str(val) not in ['-', '']]

    results = []
    for col_idx, model_name in model_cols:
        item = {'model': model_name}
        for label_raw, col_data in label_data.items():
            field = row_map.get(label_raw)
            if not field:
                continue
            val = col_data.get(col_idx)
            item[field] = cell_str(val)
        results.append(item)

    return results


# ── CAMERA PARSER ─────────────────────────────────────────────────────────────

def parse_camera_file(filepath: str, default_category: str) -> list:
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb.active

    raw_list = parse_transposed_xlsx(ws, CAMERA_ROW_MAP)
    cameras = []

    for raw in raw_list:
        model = raw.get('__model__') or raw.get('model', '')
        if not model or model == '-':
            continue

        cam = {}
        cam['model']    = model
        cam['status']   = STATUS_MAP.get(raw.get('status', '출시완료'), '출시완료')
        cam['category'] = default_category
        cam['title']    = raw.get('title', model)
        cam['resolution']      = raw.get('resolution', '-')
        cam['mp']              = parse_mp(raw.get('resolution', ''))
        cam['releaseDate']     = raw.get('releaseDate', '-')
        cam['imageSensor']     = raw.get('imageSensor', '-')

        lt, ltag = parse_lens(raw.get('lensType', '-'))
        cam['lensType']    = lt
        cam['lensTypeTag'] = ltag
        cam['focalLength'] = raw.get('focalLength', '-')
        cam['flSubtag']    = get_fl_subtag(raw.get('focalLength', ''), ltag)

        cam['fov']              = raw.get('fov', '-')
        cam['minIllumination']  = raw.get('minIllumination', '-')
        cam['wdr']              = raw.get('wdr', '-')
        cam['dayNight']         = raw.get('dayNight', '-')
        cam['compression']      = raw.get('compression', '-')
        cam['maxFrameRate']     = raw.get('maxFrameRate', '-')
        #cam['videoAnalytics']   = raw.get('videoAnalytics', '-')
        cam['analyticsTier']    = detect_analytics(raw.get('videoAnalytics',''), raw.get('keyFeature',''))
        cam['videoAnalytics'] = raw.get('videoAnalytics', '-')
        cam['analyticsFeatures'] = detect_analytics_features(
            raw.get('videoAnalytics', ''),  # Video Analytics
            raw.get('keyFeature', '')       # Key Features
        )
        cam['hasAnalytics'] = len(cam['analyticsFeatures']) > 0    
        cam['twoWayAudio']      = raw.get('twoWayAudio', '-')
        cam['audioInOut']       = raw.get('audioInOut', '-')
        cam['alarmInOut']       = raw.get('alarmInOut', '-')
        cam['edgeStorage']      = raw.get('edgeStorage', '-')
        cam['vandal']           = raw.get('vandal', '-')
        cam['outdoor']          = raw.get('outdoor', '-')
        cam['powerSource']      = raw.get('powerSource', '-')
        cam['powerConsumption'] = raw.get('powerConsumption', '-')
        cam['dimensions']       = raw.get('dimensions', '-')
        cam['weight']           = raw.get('weight', '-')
        cam['approval']         = raw.get('approval', '-')
        cam['keyFeature']       = raw.get('keyFeature', '-')
        cam['irDistance']       = raw.get('irDistance') or raw.get('__ir_domestic__', '-')
        if not cam['irDistance'] or cam['irDistance'] == '-':
            cam['irDistance'] = raw.get('__ir_domestic__', '-')

        # carry __security__ for FIPS detection
        cam['__security__'] = raw.get('__security__', '')

        detect_camera_booleans(cam)
        cam.pop('__security__', None)

        cameras.append(cam)

    return cameras


# ── RECORDER PARSER ───────────────────────────────────────────────────────────

def parse_recorder_file(filepath: str, default_series: str) -> list:
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb.active

    raw_list = parse_transposed_xlsx(ws, RECORDER_ROW_MAP)
    recorders = []

    for raw in raw_list:
        model = raw.get('__model__') or raw.get('model', '')
        if not model or model == '-':
            continue

        ch_raw = raw.get('__channels__', '0')
        m = re.search(r'(\d+)', str(ch_raw))
        ch_num = int(m.group(1)) if m else 0

        # Fallback: DR-8516 -> 16, DR-8564D -> 64, DR-2504P -> 4
        if ch_num == 0:
            m_model = re.search(r'DR-\d{2}(\d{2})', model.upper())
            if m_model:
                ch_num = int(m_model.group(1))

        title = raw.get('title', model)
        hdd = raw.get('hdd', '-')
        supported_cam = raw.get('supportedCam', 'DirectIP')
        max_bandwidth = raw.get('maxBandwidth', '-')
        rec_bandwidth = raw.get('recBandwidth', '-')
        rec_res_raw = raw.get('recRes', '-')
        video_out = raw.get('videoOut', '-')
        total_capacity_raw = raw.get('totalCapacity', '-')

        rec = {}
        rec['model']         = model
        rec['series']        = default_series
        rec['status']        = STATUS_MAP.get(raw.get('status', '출시완료'), '출시완료')
        rec['title']         = title
        rec['ch_num']        = ch_num
        rec['channels']      = str(ch_num) if ch_num else clean_channel_text(ch_raw)
        rec['maxBandwidth']  = max_bandwidth
        rec['recBandwidth']  = clean_bandwidth(rec_bandwidth)
        rec['recRes']        = clean_rec_resolution(rec_res_raw)
        rec['hdd']           = hdd
        rec['totalCapacity'] = clean_total_capacity(total_capacity_raw)
        rec['supportedCam']  = supported_cam
        rec['videoOut']      = video_out

        rec['dewarping'] = bool_cell(raw.get('__dewarping__', ''))

        rec['hasRAID'] = (
            bool_cell(raw.get('__raid__', ''))
            or text_has_raid(hdd)
        )

        rec['hasONVIF'] = (
            bool_cell(raw.get('__onvif__', ''))
            or 'onvif' in str(supported_cam or '').lower()
            or default_series != 'DR1'
        )

        rec['is4K'] = (
            bool_cell(raw.get('__4k__', ''))
            or text_has_4k(title, rec_res_raw, rec_bandwidth, video_out)
        )

        recorders.append(rec)

    return recorders

# ── MAIN ──────────────────────────────────────────────────────────────────────

def parse_all(specs_dir: str, output_dir: str):
    specs_path  = Path(specs_dir)
    output_path = Path(output_dir)
    output_path.mkdir(parents=True, exist_ok=True)

    all_cameras   = []
    all_recorders = []

    xlsx_files = sorted(specs_path.glob('*.xlsx'))
    if not xlsx_files:
        print(f"No XLSX files found in: {specs_path.resolve()}")
        sys.exit(1)

    print(f"Found {len(xlsx_files)} XLSX files in '{specs_path}'")
    print()

    for f in xlsx_files:
        name = f.name
        if is_camera_file(name):
            cat = guess_camera_category(name)
            print(f"  [CAMERA]   {name}  →  category: {cat}")
            try:
                cams = parse_camera_file(str(f), cat)
                print(f"             {len(cams)} models parsed")
                all_cameras.extend(cams)
            except Exception as e:
                print(f"             ERROR: {e}")

        elif is_recorder_file(name):
            series = guess_recorder_series(name)
            print(f"  [RECORDER] {name}  →  series: {series}")
            try:
                recs = parse_recorder_file(str(f), series)
                print(f"             {len(recs)} models parsed")
                all_recorders.extend(recs)
            except Exception as e:
                print(f"             ERROR: {e}")
        else:
            print(f"  [SKIP]     {name}")

    print()

    # Write cameras.json
    cam_path = output_path / 'cameras.json'
    with open(cam_path, 'w', encoding='utf-8') as fp:
        json.dump(all_cameras, fp, ensure_ascii=False, indent=2)
    print(f"cameras.json    → {cam_path}  ({len(all_cameras)} models)")

    # Write recorders.json
    rec_path = output_path / 'recorders.json'
    with open(rec_path, 'w', encoding='utf-8') as fp:
        json.dump(all_recorders, fp, ensure_ascii=False, indent=2)
    print(f"recorders.json  → {rec_path}  ({len(all_recorders)} models)")

    return all_cameras, all_recorders


def main():
    parser = argparse.ArgumentParser(
        description='Parse IDIS XLSX spec files → cameras.json + recorders.json'
    )
    parser.add_argument('--specs-dir',  default='specs',  help='Folder containing XLSX files (default: ./specs)')
    parser.add_argument('--output-dir', default='.',       help='Where to write JSON files (default: .)')
    args = parser.parse_args()

    parse_all(args.specs_dir, args.output_dir)
    print("\nDone.")


if __name__ == '__main__':
    main()
